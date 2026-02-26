from annotated_types import doc
from fastapi import APIRouter, Depends, HTTPException
from datetime import datetime
from bson import ObjectId
from typing import Union, List, Dict, Any
import io
import csv
import os
from openpyxl import load_workbook

from app.db.connection import db
from app.utils.auth import require_roles, hash_password
from app.models.customer_model import CustomerCreate, CustomerOut
from fastapi import Request, UploadFile, File, Form
from app.services.storage import save_customer_file
from app.services.url import build_media_url
from fastapi import Request
from app.services.storage import save_upload_file

router = APIRouter(prefix="/customer", tags=["Customer Management"])
print("Customer router loaded")

def validate_image_file(file: UploadFile):
    allowed_extensions = (".png", ".jpg", ".jpeg")
    allowed_types = ["image/png", "image/jpeg", "image/jpg"]

    filename = file.filename.lower()

    if not filename.endswith(allowed_extensions):
        raise HTTPException(
            status_code=400,
            detail=f"{file.filename} must be PNG or JPG/JPEG format"
        )

    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=400,
            detail=f"{file.filename} has invalid content type"
        )

# --------------------------------------------------------
# 🟢 BULK UPLOAD: Excel + Logo Files
# --------------------------------------------------------

@router.post("/bulk-upload")
async def bulk_upload_customers(
    excel_file: UploadFile = File(...),
    logo_files: List[UploadFile] = File(default=[]),
    linked_company_id: str = Form(None),
    user=Depends(require_roles("superadmin", "company")),
):

    print("Received logos:", [file.filename for file in logo_files])

    try:
        filename = excel_file.filename.lower()
        is_csv = filename.endswith(".csv")
        is_excel = filename.endswith((".xlsx", ".xls"))

        if not (is_csv or is_excel):
            raise HTTPException(
                status_code=400,
                detail="File must be Excel (.xlsx, .xls) or CSV (.csv)",
            )

        file_content = await excel_file.read()
        if not file_content:
            raise HTTPException(status_code=400, detail="File is empty")

        # -------- Parse File --------
        headers = []
        rows_data = []

        if is_csv:
            text_content = file_content.decode("utf-8")
            reader = csv.DictReader(io.StringIO(text_content))
            headers = [h.strip() for h in (reader.fieldnames or [])]
            rows_data = list(reader)

        else:
            workbook = load_workbook(io.BytesIO(file_content))
            worksheet = workbook.active

            headers = [str(cell.value).strip() for cell in worksheet[1]]

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        row_dict[header] = row[idx]
                rows_data.append(row_dict)

        if not headers:
            raise HTTPException(status_code=400, detail="Headers missing")

        # -------- Build Logo Map (full name + stem) --------
        logo_files_map = {}         # full filename (lowercase) -> UploadFile
        logo_files_map_no_ext = {}  # filename stem (lowercase) -> UploadFile or None if ambiguous

        for f in logo_files:
            if not f.filename:
                continue
            fname = f.filename.strip()
            key_full = fname.lower()
            logo_files_map[key_full] = f

            stem = os.path.splitext(fname)[0].lower()
            if stem in logo_files_map_no_ext:
                # mark ambiguous stems as None so we can error on ambiguity
                logo_files_map_no_ext[stem] = None
            else:
                logo_files_map_no_ext[stem] = f

        print("Logo map (full):", list(logo_files_map.keys()))
        print("Logo map (stem):", [k for k in logo_files_map_no_ext.keys()])

        # -------- Resolve Company --------
        company_id = None

        if user["role"] == "company":
            company = await db.companies.find_one({"user_id": str(user["_id"])})
            if company:
                company_id = str(company["_id"])

        elif user["role"] == "superadmin":
            if not linked_company_id:
                raise HTTPException(
                    status_code=400,
                    detail="linked_company_id required for superadmin"
                )
            company_id = linked_company_id

        # -------- Process Rows --------
        results = []
        row_num = 2
        logo_path_cache = {}

        for row_data in rows_data:

            if not any(row_data.get(f) for f in ["username", "email", "password", "full_name"]):
                row_num += 1
                continue

            try:
                logo_filename = (
                    str(row_data.get("logo_url", "")).strip()
                    or str(row_data.get("logo_file_name", "")).strip()
                )

                logo_url = None

                if logo_filename:

                        # 🔴 Case 1: Logo name given but no files uploaded
                        if not logo_files_map:
                            raise HTTPException(
                                status_code=400,
                                detail=f"Row {row_num}: Logo '{logo_filename}' mentioned but no logo files uploaded"
                            )

                        # Try exact match (case-insensitive) first, then try matching by stem (without extension)
                        logo_filename_key = str(logo_filename).strip().lower()
                        logo_file = None

                        if logo_filename_key in logo_files_map:
                            logo_file = logo_files_map[logo_filename_key]

                        else:
                            stem = os.path.splitext(logo_filename_key)[0]
                            if stem in logo_files_map_no_ext:
                                if logo_files_map_no_ext[stem] is None:
                                    # ambiguous: multiple uploaded files share the same stem
                                    raise HTTPException(
                                        status_code=400,
                                        detail=f"Row {row_num}: Ambiguous logo name '{logo_filename}' matches multiple uploaded files"
                                    )
                                logo_file = logo_files_map_no_ext[stem]

                        if not logo_file:
                            raise HTTPException(
                                status_code=400,
                                detail=f"Row {row_num}: Logo filename mismatch '{logo_filename}' not found in uploaded files"
                            )

                        validate_image_file(logo_file)

                        # Reuse cached (cache keyed by normalized name)
                        if logo_filename_key in logo_path_cache:
                            logo_url = logo_path_cache[logo_filename_key]

                        # Save if uploaded and we have a company to save under
                        elif company_id:
                            path, _ = await save_upload_file(logo_file, company_id)
                            logo_url = path   # ✅ store raw path only
                            logo_path_cache[logo_filename_key] = path

                        else:
                            print(f"Logo not uploaded: {logo_filename}")

                customer_data = {
                    "username": str(row_data.get("username")).strip(),
                    "email": str(row_data.get("email")).strip(),
                    "password": str(row_data.get("password")).strip(),
                    "full_name": str(row_data.get("full_name")).strip(),
                    "customer_company_name": row_data.get("customer_company_name"),
                    "city": row_data.get("city"),
                    "phone_number": row_data.get("phone_number"),
                    "telephone_number": row_data.get("telephone_number"),
                    "address": row_data.get("address"),
                    "linked_company_id": company_id,
                    "customer_category": str(row_data.get("customer_category")).strip()
                }

                if logo_url:
                    customer_data["logo_url"] = logo_url

                result = await create_single_customer(customer_data, user)

                results.append({
                    "success": True,
                    "row": row_num,
                    "data": result
                })

            except Exception as e:
                results.append({
                    "success": False,
                    "row": row_num,
                    "error": str(e),
                    "data": row_data
                })

            row_num += 1

        successful = sum(1 for r in results if r["success"])
        failed = len(results) - successful

        return {
            "message": "Bulk upload completed",
            "total_rows": len(results),
            "successful": successful,
            "failed": failed,
            "results": results,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def to_oid(value: str):
    try:
        return ObjectId(value)
    except:
        raise HTTPException(status_code=400, detail=f"Invalid ObjectId: {value}")

# --------------------------------------------------------
# 🟢 Helper: Validate and Create Customer Document
# --------------------------------------------------------
async def create_single_customer(data: Dict[str, Any], user: Dict):
    try:
        print("Creating single customer with data:", data)

        # -----------------------------
        # ✅ REQUIRED FIELD VALIDATION
        # -----------------------------
        required_fields = ["username", "email", "password", "full_name"]

        for field in required_fields:
            if not data.get(field):
                raise HTTPException(
                    status_code=400,
                    detail=f"{field} is required"
                )

        # -----------------------------
        # COMPANY AUTO LINK
        # -----------------------------
        if user["role"] == "company":
            company = await db.companies.find_one({"user_id": str(user["_id"])})
            if not company:
                raise HTTPException(404, "Company record not found")
            data["linked_company_id"] = str(company["_id"])

        elif user["role"] == "superadmin":
            if not data.get("linked_company_id"):
                raise HTTPException(
                    400,
                    "linked_company_id is required for superadmin"
                )

        # -----------------------------
        # DUPLICATE CHECK
        # -----------------------------
        if await db.users.find_one({"username": data["username"]}):
            raise HTTPException(400, "Username already exists")

        if await db.users.find_one({"email": data["email"]}):
            raise HTTPException(400, "Email already exists")

        # -----------------------------
        # CREATE USER
        # -----------------------------
        user_doc = {
            "username": data["username"],
            "email": data["email"],
            "password": hash_password(data["password"]),
            "role": "customer",
            "status": "active",
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow(),
        }

        inserted_user = await db.users.insert_one(user_doc)
        user_id = inserted_user.inserted_id

        # -----------------------------
        # CREATE CUSTOMER
        # -----------------------------
        customer_doc = {
            "customer_company_name": data.get("customer_company_name"),
            "full_name": data["full_name"],
            "logo_url": data.get("logo_url"),
            "city": data.get("city"),
            "phone_number": data.get("phone_number"),
            "telephone_number": data.get("telephone_number"),
            "address": data.get("address"),
            "linked_company_id": to_oid(data["linked_company_id"]),
            "customer_category": data.get("customer_category"),  
            "user_id": user_id,
            "status": "active",
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow(),
        }

        inserted = await db.customers.insert_one(customer_doc)

        return {
            "message": "Customer created successfully",
            "customer_id": str(inserted.inserted_id),
            "user_id": str(user_id),
        }

    except HTTPException:
        raise  # re-raise validation errors

    except Exception as e:
        print("🔥 INTERNAL ERROR:", str(e))
        raise HTTPException(
            status_code=500,
            detail=f"Internal error: {str(e)}"
        )

# --------------------------------------------------------
# 🟢 CREATE CUSTOMER (Single / Bulk)
# --------------------------------------------------------
@router.post("/")
async def create_customer_handler(
    request: Request,
    logo_file: UploadFile = File(None),  # optional
    linked_company_id: str = Form(None),
    user=Depends(require_roles("superadmin", "company")),
):
    content_type = request.headers.get("content-type", "")
    if user["role"] == "company":
        company = await db.companies.find_one({"user_id": str(user["_id"])})
        if company:
            company_id = str(company["_id"])

    elif user["role"] == "superadmin":
        if not linked_company_id:
            raise HTTPException(
                status_code=400,
                detail="linked_company_id required for superadmin"
            )
        company_id = linked_company_id
    # ----------------------------------
    # 🔵 CASE 1: JSON → Bulk or Single
    # ----------------------------------
    if content_type.startswith("application/json"):
        data = await request.json()

        # 🔹 Bulk
        if isinstance(data, list):
            if not data:
                raise HTTPException(status_code=400, detail="Input list cannot be empty")

            results = []
            for item in data:
                try:
                    res = await create_single_customer(item, user)
                    results.append({"success": True, "data": res})
                except HTTPException as e:
                    results.append({"success": False, "error": e.detail, "data": item})

            return {
                "message": "Bulk creation completed",
                "total": len(data),
                "results": results,
            }

        # 🔹 Single JSON (no file)
        return await create_single_customer(data, user)

    # ----------------------------------
    # 🟢 CASE 2: FORM-DATA → Single + File
    # ----------------------------------
    elif content_type.startswith("multipart/form-data"):
        form = await request.form()
        print("Form data:", form)
        # Convert form-data → dict (NO bytes issue)
        data = dict(form)

        # Remove file object from dict
        data.pop("logo_file", None)

        # Save logo if provided
        if logo_file:
            validate_image_file(logo_file)
            
            path, _ = await save_customer_file(logo_file,company_id, data.get("username"))
            data["logo_url"] = path
            print("Logo saved at:", path)
        return await create_single_customer(data, user)

    else:
        raise HTTPException(status_code=415, detail="Unsupported Content-Type")

# --------------------------------------------------------
# 🔵 LIST CUSTOMERS WITH USER + COMPANY JOIN
# --------------------------------------------------------
@router.get("/")
async def list_customers(request: Request,user=Depends(require_roles("superadmin", "company"))):

    # ✅ FILTER: Company users can only see their own customers
    match_stage = {}
    if user["role"] == "company":
        company = await db.companies.find_one({"user_id": str(user["_id"])})
        if not company:
            raise HTTPException(status_code=404, detail="Company not found for this user")
        match_stage["linked_company_id"] = company["_id"]

    pipeline = []
    
    # Add match stage only if there are filters
    if match_stage:
        pipeline.append({"$match": match_stage})
    
    pipeline.extend([
        {
            "$lookup": {
                "from": "users",
                "localField": "user_id",
                "foreignField": "_id",
                "as": "user",
            }
        },
        {"$unwind": "$user"},
        {
            "$lookup": {
                "from": "companies",
                "localField": "linked_company_id",
                "foreignField": "_id",
                "as": "company",
            }
        },
        {"$unwind": {"path": "$company", "preserveNullAndEmptyArrays": True}},
    ])

    data = []
    async for doc in db.customers.aggregate(pipeline):

        # Convert ObjectIds
        doc["id"] = str(doc.pop("_id"))
        doc["user_id"] = str(doc["user_id"])
        doc["linked_company_id"] = str(doc["linked_company_id"])

         # ✅ CUSTOMER LOGO FIX
        if doc.get("logo_url"):
            doc["logo_url"] = f"media/{doc['logo_url']}"

        doc["user"]["id"] = str(doc["user"].pop("_id"))
        doc["user"].pop("password")


        # ✅ COMPANY SAFE FIX
        # COMPANY
        if doc.get("company"):
            doc["company"]["id"] = str(doc["company"].pop("_id"))

            if doc["company"].get("logo_url") and not doc["company"]["logo_url"].startswith("media/"):
                doc["company"]["logo_url"] = f"media/{doc['company']['logo_url']}"
        else:
            doc["company"] = None

        data.append(doc)

    return data

# --------------------------------------------------------
# 🔵 GET SINGLE CUSTOMER
# --------------------------------------------------------
@router.get("/{customer_id}")
async def get_customer(customer_id: str, user=Depends(require_roles("superadmin", "company"))):

    # ✅ AUTHORIZATION: Company users can only access their own customers
    match_filter = {"_id": to_oid(customer_id)}
    
    if user["role"] == "company":
        company = await db.companies.find_one({"user_id": str(user["_id"])})
        if not company:
            raise HTTPException(status_code=404, detail="Company not found for this user")
        match_filter["linked_company_id"] = company["_id"]

    pipeline = [
        {"$match": match_filter},
        {
            "$lookup": {
                "from": "users",
                "localField": "user_id",
                "foreignField": "_id",
                "as": "user"
            }
        },
        {"$unwind": "$user"},
        {
            "$lookup": {
                "from": "companies",
                "localField": "linked_company_id",
                "foreignField": "_id",
                "as": "company",
            }
        },
        {"$unwind": {"path": "$company", "preserveNullAndEmptyArrays": True}},
    ]

    result = await db.customers.aggregate(pipeline).to_list(1)
    if not result:
        raise HTTPException(status_code=404, detail="Customer not found")

    doc = result[0]

    doc["id"] = str(doc.pop("_id"))
    doc["user_id"] = str(doc["user_id"])
    doc["linked_company_id"] = str(doc["linked_company_id"])
    # ✅ CUSTOMER LOGO FIX
    if doc.get("logo_url"):
        doc["logo_url"] = f"media/{doc['logo_url']}"

    doc["user"]["id"] = str(doc["user"].pop("_id"))
    doc["user"].pop("password")

     # ========================
    # COMPANY SAFE FIX
    # ========================
    if doc.get("company"):
        company = doc["company"]

        company["id"] = str(company.pop("_id"))

        if company.get("logo_url") and not company["logo_url"].startswith("media/"):
            company["logo_url"] = f"media/{company['logo_url']}"
    else:
        doc["company"] = None

    return doc

# --------------------------------------------------------
# 🟠 UPDATE CUSTOMER
# --------------------------------------------------------
@router.patch("/{customer_id}")
async def update_customer(
    customer_id: str,
    customer_company_name: str = Form(None),
    full_name: str = Form(None),
    city: str = Form(None),
    phone_number: str = Form(None),
    telephone_number: str = Form(None),
    address: str = Form(None),
    status: str = Form(None),
    logo_url: UploadFile = File(None),
    customer_category: str = Form(None),  # NEW FIELD
    user=Depends(require_roles("company"))
):
    # ---- Check customer exists ----
    customer = await db.customers.find_one({"_id": ObjectId(customer_id)})
    if not customer:
        raise HTTPException(404, "Customer not found")

    # ---- Build dynamic update dict ----
    update_data = {}

    if customer_company_name is not None:
        update_data["customer_company_name"] = customer_company_name
    if full_name is not None:
        update_data["full_name"] = full_name
    if city is not None:
        update_data["city"] = city
    if phone_number is not None:
        update_data["phone_number"] = phone_number
    if telephone_number is not None:
        update_data["telephone_number"] = telephone_number
    if address is not None:
        update_data["address"] = address
    if status is not None:
        update_data["status"] = status
    if logo_url is not None:
        validate_image_file(logo_url)

        from app.services.storage import save_upload_file
        path, _ = await save_upload_file(logo_url, f"customer_{customer_id}")
        update_data["logo_url"] = path
    if customer_category is not None:  # NEW FIELD
        update_data["customer_category"] = customer_category    

    if not update_data:
        raise HTTPException(400, "No fields provided to update")

    update_data["updated_at"] = datetime.utcnow()

    # ---- Update MongoDB ----
    await db.customers.update_one(
        {"_id": ObjectId(customer_id)},
        {"$set": update_data}
    )

    return {
        "message": "Customer updated successfully",
        "updated_fields": list(update_data.keys())
    }

'''@router.patch("/{customer_id}")
async def update_customer(customer_id: str, data: Dict, 
                          user=Depends(require_roles("superadmin", "company"))):

    data["updated_at"] = datetime.utcnow()

    # Can't manually change user_id
    data.pop("user_id", None)

    result = await db.customers.update_one(
        {"_id": to_oid(customer_id)},
        {"$set": data},
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")

    return {"message": "Customer updated successfully"}'''

# --------------------------------------------------------
# 🔴 DELETE CUSTOMER + LINKED USER
# --------------------------------------------------------
@router.delete("/{customer_id}")
async def delete_customer(customer_id: str, user=Depends(require_roles("superadmin", "company"))):

    customer = await db.customers.find_one({"_id": to_oid(customer_id)})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    await db.customers.delete_one({"_id": to_oid(customer_id)})
    await db.users.delete_one({"_id": customer["user_id"]})

    return {"message": "Customer and linked user deleted successfully"}
